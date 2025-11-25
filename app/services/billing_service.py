"""
Servicio de Facturación para Organizadores
Maneja la lógica de negocio del sistema de facturación
Integración con MercadoPago para sincronización de pagos
"""
from sqlalchemy.orm import Session
from fastapi import HTTPException, status
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta, timezone
from decimal import Decimal
import uuid
import mercadopago
import logging

from app.repositories.billing_repository import BillingRepository
from app.models.event import Event, EventStatus
from app.models.purchase import Purchase, PurchaseStatus
from app.models.user import User
from app.core.config import settings
from app.schemas.billing import (
    OrganizerEventBillingSchema,
    EventBillingDetailSchema,
    BillingSummarySchema,
    BillingCommissionsSchema,
    BillingCommissionSchema,
    BillingAccreditationSchema,
    PaymentMethodSchema,
    BillingTransactionSchema,
    BillingSyncResponseSchema
)
from app.utils.encryption import decrypt_mercadopago_token, encrypt_mercadopago_token
import requests

logger = logging.getLogger(__name__)


class BillingService:
    """Servicio para gestión de facturación de organizadores"""
    
    # Configuración de comisiones
    MERCADOPAGO_COMMISSION_RATE = Decimal('0.0499')  # 4.99%
    PLATFORM_COMMISSION_RATE = Decimal('0.03')  # 3%
    ACCREDITATION_DAYS = 14  # Días para acreditación
    
    def __init__(self, db: Session):
        self.db = db
        self.repository = BillingRepository(db)
        self.mp_sdk = None  # Se inicializa cuando sea necesario
    
    # ============================================================
    # Servicios Públicos
    # ============================================================
    
    def get_organizer_events(self, organizer_id: uuid.UUID) -> List[OrganizerEventBillingSchema]:
        """
        Obtener lista de eventos del organizador con datos de facturación
        """
        events = self.repository.get_organizer_events(organizer_id)
        
        result = []
        for event in events:
            billing_data = self._calculate_event_billing_summary(event)
            result.append(
                OrganizerEventBillingSchema(
                    id=str(event.id),
                    title=event.title,
                    startDate=event.startDate,
                    totalRevenue=billing_data['total_revenue'],
                    totalTransactions=billing_data['total_transactions'],
                    netAmount=billing_data['net_amount'],
                    status=event.status.value
                )
            )
        
        return result
    
    def get_event_billing_detail(
        self, 
        event_id: uuid.UUID, 
        organizer_id: uuid.UUID
    ) -> EventBillingDetailSchema:
        """
        Obtener detalle completo de facturación de un evento
        """
        # Verificar que el evento pertenezca al organizador
        event = self.repository.get_event_by_id(event_id, organizer_id)
        
        if not event:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Evento no encontrado o no tienes permisos para verlo"
            )
        
        # Calcular resumen
        summary = self._build_billing_summary(event)
        
        # Calcular distribución de métodos de pago
        payment_methods = self._build_payment_methods_distribution(event)
        
        # Construir lista de transacciones
        transactions = self._build_transactions_list(event)
        
        # Última sincronización (usar la fecha más reciente de las compras)
        last_sync = datetime.now(timezone.utc)
        if event.purchases:
            purchase_dates = [p.updated_at for p in event.purchases if p.updated_at]
            if purchase_dates:
                last_sync = max(purchase_dates)
        
        return EventBillingDetailSchema(
            eventId=str(event.id),
            eventName=event.title,
            eventDate=event.startDate,
            summary=summary,
            paymentMethods=payment_methods,
            transactions=transactions,
            lastSync=last_sync
        )
    
    def sync_event_billing(
        self, 
        event_id: uuid.UUID, 
        organizer_id: uuid.UUID
    ) -> BillingSyncResponseSchema:
        """
        Sincronizar datos de facturación con MercadoPago
        """
        # Verificar que el evento pertenezca al organizador
        event = self.repository.get_event_by_id(event_id, organizer_id)
        
        if not event:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Evento no encontrado"
            )
        
        # Obtener el organizador para usar su token de MP
        organizer = self.db.query(User).filter(User.id == organizer_id).first()
        
        if not organizer or not organizer.isMercadopagoConnected:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No tienes una cuenta de MercadoPago vinculada"
            )
        
        # Inicializar SDK con el token del organizador
        self._initialize_mp_sdk(organizer.mercadopagoAccessToken)
        
        # Sincronizar todas las compras del evento
        transactions_updated = 0
        purchases = self.repository.get_event_purchases(event_id)
        
        for purchase in purchases:
            if purchase.mercadopago_preference_id:
                try:
                    # Consultar estado del pago en MercadoPago
                    payment_info = self._get_payment_from_mercadopago(
                        purchase.mercadopago_preference_id
                    )
                    
                    if payment_info:
                        # Actualizar la compra con los datos de MP
                        self.repository.update_purchase_from_mercadopago(
                            purchase.id,
                            payment_info
                        )
                        transactions_updated += 1
                        
                except Exception as e:
                    logger.error(f"Error sincronizando compra {purchase.id}: {str(e)}")
                    continue
        
        return BillingSyncResponseSchema(
            message=f"Sincronización completada. {transactions_updated} transacciones actualizadas.",
            transactionsUpdated=transactions_updated,
            lastSync=datetime.now(timezone.utc)
        )
    
    # ============================================================
    # Generación de Reportes
    # ============================================================
    
    def generate_pdf_report(
        self, 
        event_id: uuid.UUID, 
        organizer_id: uuid.UUID
    ) -> bytes:
        """
        Generar reporte PDF de facturación
        """
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from io import BytesIO
        
        # Obtener datos del evento
        detail = self.get_event_billing_detail(event_id, organizer_id)
        
        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph(f"<b>Reporte de Facturación - {detail.eventName}</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Información del evento
        event_info = Paragraph(f"Fecha del Evento: {detail.eventDate.strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
        story.append(event_info)
        story.append(Spacer(1, 0.2*inch))
        
        # Resumen financiero
        summary_data = [
            ['Métrica', 'Valor'],
            ['Ingresos Totales', f"S/. {detail.summary.totalRevenue:.2f}"],
            ['Comisión MercadoPago', f"S/. {detail.summary.commissions.mercadoPago.amount:.2f}"],
            ['Comisión Plataforma', f"S/. {detail.summary.commissions.platform.amount:.2f}"],
            ['Total Comisiones', f"S/. {detail.summary.commissions.total:.2f}"],
            ['Monto Neto', f"S/. {detail.summary.netAmount:.2f}"],
            ['Transacciones', str(detail.summary.totalTransactions)],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Métodos de pago
        if detail.paymentMethods:
            methods_title = Paragraph("<b>Distribución de Métodos de Pago</b>", styles['Heading2'])
            story.append(methods_title)
            story.append(Spacer(1, 0.1*inch))
            
            methods_data = [['Método', 'Cantidad', 'Monto', 'Porcentaje']]
            for method in detail.paymentMethods:
                methods_data.append([
                    method.method,
                    str(method.count),
                    f"S/. {method.amount:.2f}",
                    f"{method.percentage:.1f}%"
                ])
            
            methods_table = Table(methods_data)
            methods_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(methods_table)
        
        # Generar PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer.read()
    
    def generate_excel_report(
        self, 
        event_id: uuid.UUID, 
        organizer_id: uuid.UUID
    ) -> bytes:
        """
        Generar reporte Excel de facturación
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Obtener datos del evento
        detail = self.get_event_billing_detail(event_id, organizer_id)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturación"
        
        # Título
        ws['A1'] = f"Reporte de Facturación - {detail.eventName}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Información del evento
        ws['A3'] = "Fecha del Evento:"
        ws['B3'] = detail.eventDate.strftime('%d/%m/%Y %H:%M')
        
        # Resumen financiero
        ws['A5'] = "RESUMEN FINANCIERO"
        ws['A5'].font = Font(bold=True)
        
        ws['A6'] = "Ingresos Totales"
        ws['B6'] = float(detail.summary.totalRevenue)
        ws['B6'].number_format = '"S/." #,##0.00'
        
        ws['A7'] = "Comisión MercadoPago"
        ws['B7'] = float(detail.summary.commissions.mercadoPago.amount)
        ws['B7'].number_format = '"S/." #,##0.00'
        
        ws['A8'] = "Comisión Plataforma"
        ws['B8'] = float(detail.summary.commissions.platform.amount)
        ws['B8'].number_format = '"S/." #,##0.00'
        
        ws['A9'] = "Total Comisiones"
        ws['B9'] = float(detail.summary.commissions.total)
        ws['B9'].number_format = '"S/." #,##0.00'
        
        ws['A10'] = "Monto Neto"
        ws['B10'] = float(detail.summary.netAmount)
        ws['B10'].number_format = '"S/." #,##0.00'
        ws['B10'].font = Font(bold=True)
        
        # Transacciones
        ws['A12'] = "TRANSACCIONES"
        ws['A12'].font = Font(bold=True)
        
        headers = ['Fecha', 'Email', 'Monto', 'Com. MP', 'Com. Plat', 'Neto', 'Estado', 'Método']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=13, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        row = 14
        for transaction in detail.transactions:
            ws.cell(row=row, column=1, value=transaction.date.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=2, value=transaction.buyerEmail)
            ws.cell(row=row, column=3, value=float(transaction.amount))
            ws.cell(row=row, column=4, value=float(transaction.mpCommission))
            ws.cell(row=row, column=5, value=float(transaction.platformCommission))
            ws.cell(row=row, column=6, value=float(transaction.netAmount))
            ws.cell(row=row, column=7, value=transaction.status)
            ws.cell(row=row, column=8, value=transaction.paymentMethod)
            
            # Formato de moneda
            for col in [3, 4, 5, 6]:
                ws.cell(row=row, column=col).number_format = '"S/." #,##0.00'
            
            row += 1
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.read()
    
    # ============================================================
    # Métodos Privados de Cálculo
    # ============================================================
    
    def _calculate_event_billing_summary(self, event: Event) -> Dict[str, Any]:
        """Calcular resumen de facturación de un evento"""
        revenue_data = self.repository.calculate_event_revenue(event.id)
        commissions_data = self.repository.calculate_event_commissions(event.id)
        
        total_revenue = revenue_data['total_revenue']
        total_commission = commissions_data['total_commission']
        
        return {
            'total_revenue': total_revenue,
            'total_transactions': revenue_data['total_transactions'],
            'net_amount': total_revenue - total_commission
        }
    
    def _build_billing_summary(self, event: Event) -> BillingSummarySchema:
        """Construir esquema de resumen de facturación"""
        revenue_data = self.repository.calculate_event_revenue(event.id)
        commissions_data = self.repository.calculate_event_commissions(event.id)
        accreditation_data = self.repository.calculate_accreditation_status(event.id)
        
        total_revenue = revenue_data['total_revenue']
        mp_commission = commissions_data['mp_commission']
        platform_commission = commissions_data['platform_commission']
        total_commission = commissions_data['total_commission']
        
        return BillingSummarySchema(
            totalRevenue=total_revenue,
            totalTransactions=revenue_data['total_transactions'],
            commissions=BillingCommissionsSchema(
                mercadoPago=BillingCommissionSchema(
                    amount=mp_commission,
                    percentage=self.MERCADOPAGO_COMMISSION_RATE * 100
                ),
                platform=BillingCommissionSchema(
                    amount=platform_commission,
                    percentage=self.PLATFORM_COMMISSION_RATE * 100
                ),
                total=total_commission
            ),
            netAmount=total_revenue - total_commission,
            accreditation=BillingAccreditationSchema(
                credited=accreditation_data['credited'],
                pending=accreditation_data['pending'],
                nextDate=accreditation_data['next_date']
            )
        )
    
    def _build_payment_methods_distribution(self, event: Event) -> List[PaymentMethodSchema]:
        """Construir distribución de métodos de pago"""
        methods_data = self.repository.calculate_payment_methods_distribution(event.id)
        
        return [
            PaymentMethodSchema(
                method=method['method'],
                count=method['count'],
                amount=method['amount'],
                percentage=method['percentage']
            )
            for method in methods_data
        ]
    
    def _build_transactions_list(self, event: Event) -> List[BillingTransactionSchema]:
        """Construir lista de transacciones"""
        purchases = self.repository.get_completed_purchases(event.id)
        
        transactions = []
        for purchase in purchases:
            mp_commission = purchase.total_amount * self.MERCADOPAGO_COMMISSION_RATE
            platform_commission = purchase.total_amount * self.PLATFORM_COMMISSION_RATE
            net_amount = purchase.total_amount - mp_commission - platform_commission
            
            # Calcular fecha de acreditación
            accreditation_date = None
            if purchase.payment_date:
                accreditation_date = purchase.payment_date + timedelta(days=self.ACCREDITATION_DAYS)
            
            # Link a MercadoPago
            mp_link = None
            if purchase.payment_reference:
                mp_link = f"https://www.mercadopago.com.pe/activities/{purchase.payment_reference}"
            
            transactions.append(
                BillingTransactionSchema(
                    id=str(purchase.id),
                    mpPaymentId=purchase.payment_reference,
                    date=purchase.created_at,
                    buyerEmail=purchase.buyer_email,
                    amount=purchase.total_amount,
                    mpCommission=mp_commission,
                    platformCommission=platform_commission,
                    netAmount=net_amount,
                    status=self._map_purchase_status(purchase.status),
                    paymentMethod=purchase.payment_method.value if purchase.payment_method else 'UNKNOWN',
                    accreditationDate=accreditation_date,
                    mpLink=mp_link
                )
            )
        
        return transactions
    
    def _map_purchase_status(self, status: PurchaseStatus) -> str:
        """Mapear estado de compra a estado de transacción"""
        status_map = {
            PurchaseStatus.COMPLETED: 'approved',
            PurchaseStatus.PENDING: 'pending',
            PurchaseStatus.REJECTED: 'rejected',
            PurchaseStatus.REFUNDED: 'refunded',
            PurchaseStatus.FAILED: 'rejected',
            PurchaseStatus.CANCELLED: 'rejected'
        }
        return status_map.get(status, 'pending')
    
    # ============================================================
    # Integración con MercadoPago
    # ============================================================
    
    def _initialize_mp_sdk(self, access_token: str):
        """Inicializar SDK de MercadoPago con token del organizador"""
        # Desencriptar el token si está encriptado
        try:
            decrypted_token = decrypt_mercadopago_token(access_token)
            self.mp_sdk = mercadopago.SDK(decrypted_token)
        except:
            # Si falla la desencriptación, asumir que está en texto plano
            logger.warning("Token no encriptado. Considerar migrar a formato encriptado.")
            self.mp_sdk = mercadopago.SDK(access_token)
    
    def _get_payment_from_mercadopago(self, preference_id: str) -> Optional[Dict[str, Any]]:
        """
        Consultar información de un pago en MercadoPago
        """
        if not self.mp_sdk:
            logger.error("SDK de MercadoPago no inicializado")
            return None
        
        try:
            # Buscar pagos asociados a la preferencia
            filters = {
                "external_reference": preference_id
            }
            
            response = self.mp_sdk.payment().search(filters=filters)
            
            if response["status"] == 200 and response["response"]["results"]:
                payment_data = response["response"]["results"][0]
                return {
                    'id': str(payment_data.get('id')),
                    'status': payment_data.get('status'),
                    'status_detail': payment_data.get('status_detail'),
                    'payment_method_id': payment_data.get('payment_method_id'),
                    'date_approved': payment_data.get('date_approved'),
                    'date_created': payment_data.get('date_created'),
                    'transaction_amount': payment_data.get('transaction_amount'),
                    'fee_details': payment_data.get('fee_details', [])
                }
            
            return None
            
        except Exception as e:
            logger.error(f"Error consultando pago en MercadoPago: {str(e)}")
            return None
    
    # ============================================================
    # Gestión de Tokens
    # ============================================================
    
    def refresh_mercadopago_token(self, user_id: uuid.UUID) -> bool:
        """
        Renovar el token de acceso de MercadoPago usando el refresh token
        
        MercadoPago tokens expiran después de 6 meses.
        Este método debe ser llamado antes de que expire.
        
        Returns:
            True si se renovó exitosamente, False si falló
        """
        try:
            # Obtener usuario
            user = self.db.query(User).filter(User.id == user_id).first()
            
            if not user or not user.mercadopagoRefreshToken:
                logger.error(f"Usuario {user_id} sin refresh token")
                return False
            
            # Desencriptar refresh token
            try:
                refresh_token = decrypt_mercadopago_token(user.mercadopagoRefreshToken)
            except:
                # Token no encriptado
                refresh_token = user.mercadopagoRefreshToken
            
            # Llamar al endpoint de refresh de MercadoPago
            url = "https://api.mercadopago.com/oauth/token"
            
            data = {
                "grant_type": "refresh_token",
                "client_id": settings.MERCADOPAGO_CLIENT_ID,
                "client_secret": settings.MERCADOPAGO_CLIENT_SECRET,
                "refresh_token": refresh_token
            }
            
            response = requests.post(url, json=data)
            
            if response.status_code != 200:
                logger.error(f"Error renovando token MP: {response.text}")
                return False
            
            token_data = response.json()
            
            # Actualizar tokens en la base de datos (encriptados)
            user.mercadopagoAccessToken = encrypt_mercadopago_token(
                token_data['access_token']
            )
            
            if 'refresh_token' in token_data:
                user.mercadopagoRefreshToken = encrypt_mercadopago_token(
                    token_data['refresh_token']
                )
            
            # Actualizar fecha de expiración
            if 'expires_in' in token_data:
                from datetime import timedelta
                user.mercadopagoTokenExpires = datetime.now(timezone.utc) + timedelta(
                    seconds=token_data['expires_in']
                )
            
            self.db.commit()
            
            logger.info(f"✅ Token renovado exitosamente para usuario {user.email}")
            return True
            
        except Exception as e:
            logger.error(f"Error renovando token de MercadoPago: {str(e)}")
            self.db.rollback()
            return False
    
    def check_and_refresh_token_if_needed(self, user_id: uuid.UUID) -> bool:
        """
        Verificar si el token está próximo a expirar y renovarlo automáticamente
        
        Verifica si el token expira en menos de 7 días y lo renueva.
        
        Returns:
            True si el token es válido o se renovó exitosamente
        """
        try:
            user = self.db.query(User).filter(User.id == user_id).first()
            
            if not user or not user.isMercadopagoConnected:
                return False
            
            # Si no hay fecha de expiración, asumir que está válido
            if not user.mercadopagoTokenExpires:
                return True
            
            # Verificar si expira en menos de 7 días
            days_until_expiry = (user.mercadopagoTokenExpires - datetime.now(timezone.utc)).days
            
            if days_until_expiry < 7:
                logger.info(f"Token expira en {days_until_expiry} días. Renovando...")
                return self.refresh_mercadopago_token(user_id)
            
            return True
            
        except Exception as e:
            logger.error(f"Error verificando expiración de token: {str(e)}")
            return False
